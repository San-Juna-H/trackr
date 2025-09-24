import rumps
import threading
import time
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook

TIMER_MINUTES = 25
XLSX_FILE = Path.home() / "Desktop" / "timesheet.xlsx"
ICON_PATH = "trackr.icns"

def notify(title, message):
    rumps.notification(title, "", message)

def append_to_xlsx(task, tag, start_dt, end_dt, duration_str):
    if XLSX_FILE.exists():
        wb = load_workbook(XLSX_FILE)
        ws = wb["Timesheet"] if "Timesheet" in wb.sheetnames else wb.active
        ws.title = "Timesheet"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Timesheet"
        ws.append(["Task", "Tag", "Start Date", "Start Time", "End Time", "Duration (MM:SS)"])

    ws.append([
        task,
        tag,
        start_dt.strftime("%Y-%m-%d"),
        start_dt.strftime("%H:%M:%S"),
        end_dt.strftime("%H:%M:%S"),
        duration_str
    ])
    wb.save(XLSX_FILE)

class TrackrTimer(threading.Thread):
    def __init__(self, task, tag, app):
        super().__init__()
        self.task = task
        self.tag = tag
        self.app = app
        self.remaining = TIMER_MINUTES * 60
        self.running = True
        self.start_time = datetime.now()

    def run(self):
        while self.running and self.remaining > 0:
            mins, secs = divmod(self.remaining, 60)
            self.app.title = f"⏱️ {mins:02d}:{secs:02d}"
            time.sleep(1)
            self.remaining -= 1

        end_time = datetime.now()
        duration_sec = int((end_time - self.start_time).total_seconds())
        duration_str = f"{duration_sec // 60:02d}:{duration_sec % 60:02d}"
        append_to_xlsx(self.task, self.tag, self.start_time, end_time, duration_str)

        status = "finished" if self.running else "stopped early"
        notify("Timer Complete" if self.running else "Timer Stopped",
               f"'{self.task}' [{self.tag}] {status}.")
        self.app.title = "trackr"
        rumps.quit_application()

    def stop(self):
        self.running = False

class TrackrApp(rumps.App):
    def __init__(self):
        super().__init__("trackr", quit_button=None, icon=ICON_PATH)
        task_input = rumps.Window("What are you going to work on?", "Task", ok="OK").run()
        tag_input = rumps.Window("Add a tag or category for this task.", "Tag", ok="OK").run()
        if not task_input.text or not tag_input.text:
            rumps.alert("Task or Tag not entered. Exiting app.")
            rumps.quit_application()
        self.timer = TrackrTimer(task_input.text, tag_input.text, self)
        self.timer.start()
        self.menu = ["Stop Early"]

    @rumps.clicked("Stop Early")
    def stop_timer(self, _):
        self.timer.stop()

if __name__ == "__main__":
    TrackrApp().run()
